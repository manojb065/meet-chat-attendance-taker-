import openpyxl as op
from openpyxl import load_workbook
import re
import os

pattern=re.compile(r'1[a-zA-Z]{2}[1-9]{2}[a-zA-Z]{2}[0-9]{3}')
pattern2=re.compile(r':\s.*')
xpath=input("enter your xlsx file path ")
xfile=input("enter your attendence folder path ")

day=int(input("enter day "))

col=6+day

wb = load_workbook(filename = xpath)

names = wb.sheetnames
sheet1=wb[names[0]]
sheet2=wb[names[2]]

dict2={}



for row in sheet1.iter_cols(min_row=2,min_col=4,max_col=4):
    for cell in row:
        if cell.value!=None:
            dict2[cell.value.lower()]=0
        


for i in  os.listdir(xfile):
    file =open(os.path.join(xfile,i),"r")
    read=file.readline()
    while read!="":
        temp=re.findall(pattern2,read)
        if temp!=[]:
            read=temp[0]
            temp=re.findall(pattern,read)
            if temp!=[]:
                if temp[0].lower() in dict2 :
                    dict2[temp[0].lower()]+=1 
                else:
                    dict2[temp[0].lower()]=1
        read=file.readline()
    

    
for row in sheet1.iter_cols(min_row=2,min_col=4,max_col=4):
    for cell in row:
        data=sheet1.cell(column=col,row=cell.row)
        if cell.value!=None and cell.value.lower() in dict2:
            if dict2[cell.value.lower()] > 2:
                data.value=2
            elif dict2[cell.value.lower()]==0:
                data.value=0
            else:
                data.value=1
            
    
wb.save(xpath)
file.close()
print("done 👍")
